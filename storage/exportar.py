from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def exportar_resultados(resultados, nombre_archivo="resultados_armado_pcs.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  

    for perfil in resultados:
        nombre_hoja = perfil["nombre"][:31] 
        ws = wb.create_sheet(title=nombre_hoja)

        fila = 1

        # ===== ENCABEZADO PERFIL =====
        ws[f"A{fila}"] = f"Nombre del perfil: {perfil['nombre']}"
        ws[f"A{fila}"].font = Font(bold=True)
        fila += 1

        ws[f"A{fila}"] = f"Tipo de perfil: {perfil['tipo']}"
        fila += 1

        ws[f"A{fila}"] = f"Presupuesto total: ${perfil['presupuesto']}"
        fila += 2

        # ===== PCs ARMADAS =====
        for idx, pc in enumerate(perfil["pcs"], start=1):
            ws[f"A{fila}"] = f"PC {idx}"
            ws[f"A{fila}"].font = Font(bold=True)
            fila += 1

            for tipo, componente in pc.componentes.items():
                ws[f"A{fila}"] = tipo.upper()
                ws[f"B{fila}"] = componente["nombre"]
                ws[f"C{fila}"] = componente["precio"]
                ws[f"C{fila}"].number_format = '"$"#.##0'
                fila += 1

            fila += 1
            ws[f"B{fila}"] = "Total gastado:"
            ws[f"C{fila}"] = pc.gastado
            ws[f"C{fila}"].number_format = '"$"#.##0'
            ws[f"B{fila}"].font = Font(bold=True)

            fila += 1
            ws[f"B{fila}"] = "Presupuesto restante:"
            ws[f"C{fila}"] = pc.presupuesto_restante()
            ws[f"C{fila}"].number_format = '"$"#.##0'
            ws[f"B{fila}"].font = Font(bold=True)

            fila += 1
            ws[f"B{fila}"] = "Presupuesto utilizado (%):"
            ws[f"C{fila}"] = (pc.gastado / perfil['presupuesto'])
            ws[f"B{fila}"].font = Font(bold=True)
            ws[f"C{fila}"].number_format = '0.00%'


            fila += 3  

        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 35

    wb.save(nombre_archivo)
